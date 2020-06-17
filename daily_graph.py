import ast
import datetime
import json
import math
import os
import time
import traceback

import bs4
import openpyxl
import PIL
from PIL import Image, ImageDraw, ImageFont
import requests
import schedule

os.chdir(os.path.dirname(os.path.abspath(__file__))) #カレントディレクトリを同一階層に(これをやらないとc:\users\username内にファイルを作りやがる)

def unexpected_error():
    """
    予期せぬエラーが起きたときの対処
    エラーメッセージ全文と発生時刻をウェブフックで通知"""

    now = datetime.datetime.now().strftime("%H:%M") #今何時？
    error_msg = f"```\n{traceback.format_exc()}```" #エラーメッセージ全文(ここに絶対パスが載ります、もしユーザーフォルダ名が本名だと大変なことに・・・)
    #webhookで投稿する中身
    main_content = {
        "username": "ERROR", #表示されるwebhook名
        "avatar_url": "https://cdn.discordapp.com/attachments/644880761081561111/703088291066675261/warning.png", #使用アイコン
        "content": "<@523303776120209408>", #けいにメンション
        "embeds": [ #エラー内容・発生時間まとめ
            {
                "title": "エラーが発生しました",
                "description": error_msg,
                "color": 0xff0000,
                "footer": {
                    "text": now
                }
            }
        ]
    }
    break_webhook_url = "https://discordapp.com/api/webhooks/以下略"
    requests.post(break_webhook_url, json.dumps(main_content), headers={'Content-Type': 'application/json'}) #エラーメッセをウェブフックに投稿


def pyxl():
    """
    jsonデータをexcelに書き込んでいく"""

    try:
        today = datetime.date.today().strftime(r"%Y%m%d") #str型の今日の日付

        workbook = openpyxl.Workbook() #新しいワークブックを作成
        worksheet = workbook.worksheets[0] #1つ目のワークシートを指定

        #元データ取得
        with open(f"{today}.json", mode="r") as f:
            player_data_dict = json.load(f)

        #横軸(時間)書き込み
        for h in range(24):
            cell = worksheet.cell(1, h+2)
            cell.value = f"{h}時"
        cell = worksheet.cell(1, 26)
        cell.value = "23:58最終"

        #データからMCIDと整地量を取得しエクセルに書き込んでいく
        i = 2
        for mcid in player_data_dict:
            cell = worksheet.cell(i, 1) #i行、A列のセルを指定
            cell.value = mcid #Aiの内容をmcidにする
            player_data = player_data_dict[mcid] #mcidの時間毎の整地量がまとめられた辞書
            j = 2
            #時間毎の整地量を書き込んでいく
            for hour in player_data:
                raw_data = player_data[hour] #時間毎の整地量
                cell = worksheet.cell(i, j) #i行、j列のセルを指定(※openpyxlモジュールではA列を1としZ列を26、AA列を27以下略とする)
                cell.value = raw_data #jiの内容を時間毎の整地量にする
                j += 1 #次の時間へ
            i += 1 #次のmcidへ

        workbook.save(f"{today}.xlsx") #xlsxファイルとして保存

    except:
        unexpected_error() #こちらでは考えつかないエラーが起きたときunexpected_error()に回す。


def draw_graph():
    """
    JSONファイルからデータを抜き出し、グラフ化して画像出力する"""

    try:
        today = datetime.date.today().strftime(r"%Y%m%d") #str型の今日の日付

        with open(f"{today}.json", mode="r") as f:
            player_data_dict = json.load(f)

        for mcid in player_data_dict:
            max_data = player_data_dict[mcid]["23_58"] #整地量の最大値を取得。これが画像の縦のpx数を決める
            break

        px = 320 #pxの最小値は150
        #最大整地量に応じた規定の大きさになるまで拡大
        while True:
            if (max_data/50000) <= px:
                break
            else:
                px += 100
        px += 50

        haikei = Image.new(mode="RGB", size=(1550, px), color=0x000000)

        font = ImageFont.truetype(r"c:\Windows\Fonts\UDDigiKyokashoN-R.ttc",size=16) #使用フォントを指定。Windows標準の位置にないとバグる

        hour = ImageDraw.Draw(haikei)
        h_place = 50

        #横軸(時間)を書く
        for h in range(24):
            hour.text((h_place, px-30), text=f"{h}時", font=font, fill=0xffffff)
            h_place += 50
        hour.text((h_place, px-30), text="最終", font=font, fill=0xffffff)
        
        #画像範囲内に目盛り線を引く
        scale = ImageDraw.Draw(haikei)
        i = 0
        while True:
            scale.line((60, px-35-i, 1250, px-35-i), fill=0xaaaaaa, width=1)
            scale.text((10, px-35-i-8), text=f"{i*5}万", font=font, fill=0xffffff)
            i += 100
            if i > px:
                break
        
        #順位ごとに違う色付け
        color_list = [
            0x00d2f4,
            0x777777,
            0x3c65bc,
            0x0000ff,
            0x00ff00,
            0xff0000,
            0xffff00,
            0xff00ff,
            0x00ffff,
            0x770000,
            0x007700,
            0x000077,
            0x777700,
            0x770077,
            0x007777,
            0x0077ff,
            0x7700ff,
            0x77ff00,
            0xff0077,
            0xff7700
        ]
        
        #上位20人の時間毎の整地量を点として入力&MCIDと最終整地量入力
        i = 0
        wide = math.floor(px/20)
        place = math.floor(wide/2) - 8
        data = ImageDraw.Draw(haikei)
        xy_list_list = [] #線を引くときに使うのでその準備
        mcid_break_place_list = [] #最終点とMCIDを紐づけするのに使うのでその準備
        for mcid in player_data_dict:
            break_amount = player_data_dict[mcid]["23_58"] #最終整地量を取得
            break_amount = "{:,}".format(break_amount)
            data.text((1340, place), text=f"{mcid}: {break_amount}", fill=0xffffff, font=font) #1位から順に右上から等間隔にMCIDと最終整地量を入力
            mcid_break_place_list.append(place+8)
            place += wide

            #時間毎の点を入力
            d_x = 50 + 10
            xy_list = [] #線を引くときに使うのでその準備
            for hour in player_data_dict[mcid]:
                raw_data = player_data_dict[mcid][hour] #時間毎の整地量を取得
                d_y = math.floor(raw_data/50000) #座標を決定
                xy = (d_x, px-35-d_y) #線を引くときに使うのでその準備
                xy_list.append(xy) #線を引くときに使うのでその準備
                data.ellipse((d_x-3, px-35-d_y-3, d_x+3, px-35-d_y+3), fill=color_list[i]) #点を打つ
                d_x += 50

            xy_list_list.append(xy_list) #線を引くときに使うのでその準備
            i += 1

        last_break_place_list = []
        #点と点を結ぶ線を入力
        for xy_list in xy_list_list:
            for i in range(24):
                x, y = xy_list[i]
                x_, y_ = xy_list[i+1]
                data.line((x, y, x_, y_), fill=color_list[xy_list_list.index(xy_list)], width=4) #点と点を結ぶ線を引く
            last_break_place_list.append((x_, y_))

        for xy_tuple in last_break_place_list:
            x, y = xy_tuple
            i = last_break_place_list.index(xy_tuple)
            data.line((x, y, 1340, mcid_break_place_list[i]), fill=color_list[i], width=2) #最終整地量点とMCIDを結ぶ線を引く

        haikei.save(f"{today}.png") #本日のグラフとしてpng形式で保存

    except:
        unexpected_error() #こちらでは考えつかないエラーが起きたときunexpected_error()に回す。


def push_files():
    """
    日間上位20人の時間毎の整地量をまとめたJSONファイルとxlsxファイルをウェブフックを使い投稿する"""

    try:
        today = datetime.date.today().strftime(r"%Y%m%d") #str型の今日の日付
        with open(f"{today}.json", mode="r") as f1: #本日の上位20人のデータをまとめたJSONファイルを指定
            with open(f"{today}.xlsx",mode="rb") as f2: #本日の上位20人のデータをまとめたxlsxファイルを指定
                with open(f"{today}.png", mode="rb") as f3: #本日の上位20人のグラフのpngファイルを指定
                    files = { #指定したファイル達
                        "file1": f1,
                        "file2": f2,
                        "file3": f3
                    }
                    break_webhook_url = "https://discordapp.com/api/webhooks/以下略"
                    requests.post(break_webhook_url, files=files) #JSONファイルをウェブフックに投稿

        os.remove(f"{today}.json") #ファイルが増え続けていくので送信したら消す
        os.remove(f"{today}.xlsx") #同上
        os.remove(f"{today}.png") #同上

    except:
        unexpected_error()


def login():
    """
    プログラム起動時の処理"""

    login_webhook_url = "https://discordapp.com/api/webhooks/以下略"
    main_content = {
            "content": "プログラムが起動しました"
        }
    requests.post(login_webhook_url, main_content) #ログインメッセをウェブフックに投稿
    print("起動しました")


def job_every_hour():
    """
    毎時の処理
    整地鯖ランキングページにアクセスし、全員のMCID・整地量を取得。
    player_data.jsonとして保存する"""

    try:
        #player_data.jsonが見つからなかった時の対処
        try:
            with open("player_data.json", mode="r") as f:
                player_data_dict = json.load(f)
        except FileNotFoundError:
            player_data_dict = {}

        #KeyErrorが出れば異常なし。出なかった場合以前のデータ取得に失敗している
        try:
            if player_data_dict["today"] == "Error":
                pass
        except KeyError:
        
            #整地鯖ランキングページにアクセス&データを記録する
            i = 0
            while True:
                url = f"https://w4.minecraftserver.jp/api/ranking?type=break&offset={i*20}&lim=20&duration=daily"
                try:
                    res = requests.get(url) #整地鯖ランキングページにアクセス
                    res.raise_for_status() #これは自分でもよくわからん。書いておかないとおかしくなる
                    player_data_every_hour_dict = ast.literal_eval(str(bs4.BeautifulSoup(res.text, "html.parser"))) #全体を辞書型で取得
                    player_data_every_hour_list = player_data_every_hour_dict["ranks"] #必要な部分のみリスト型で取得

                    for player_data_every_hour in player_data_every_hour_list: #リストから一つずつデータを抜き出す
                        mcid = player_data_every_hour["player"]["name"] #リストからMCIDを取得
                        raw_data = player_data_every_hour["data"]["raw_data"] #リストから整地量を取得
                        now = datetime.datetime.now() #今何時？
                        #以前からplayer_dataがあれば追記、なければ新規作成
                        try:
                            player_data = player_data_dict[mcid]
                        except KeyError:
                            player_data = {}
                        player_data[f"{now.hour}"] = int(raw_data) #プレイヤーの時間毎の整地量が格納される辞書
                        player_data_dict[mcid] = player_data #上記の辞書を全MCIDぶん格納する辞書

                    #まだ続きがありそうなら続ける、なければループを抜ける
                    if len(player_data_every_hour_list) == 20:
                        i += 1
                    else:
                        break
                #ウェブページに不調があったときの処理
                except requests.exceptions.HTTPError:
                    main_content = {
                        "content": "今日は無理だな(確信)"
                    }
                    break_webhook_url = "https://discordapp.com/api/webhooks/以下略"
                    requests.post(break_webhook_url, main_content) #諦めメッセをウェブフックに投稿
                    player_data_dict = {"today":"Error"} #辞書の中身をエラー発生時用に変更、こうなってると70行目で省かれてexcept内は実行されない
                    break

            #作成したplayer_data_dictをjson形式に変換して保存
            with open("player_data.json", mode="w") as f: #書き込みモードで開く
                player_data_json = json.dumps(player_data_dict, indent=4) #JSON形式に変換
                f.write(player_data_json) #書く

    except:
        unexpected_error() #こちらでは考えつかないエラーが起きたときunexpected_error()に回す。


def job_every_day():
    """
    毎日の処理
    毎日23時58分に整地鯖ランキングページにアクセスし、上位20人のMCID・整地量・順位を取得。"""

    try:
        break_webhook_url = "https://discordapp.com/api/webhooks/以下略"
        now = datetime.datetime.now() #datetime型の現在時刻
        #そもそもファイルなしを省く
        try:
            with open("player_data.json", mode="r") as f:
                player_data_dict = json.load(f)
        except FileNotFoundError:
            main_content = {
                "content": f"{now.month}月{now.day}日 : データなし"
            }
            requests.post(break_webhook_url, main_content) #結果をウェブフックに投稿
            return

        #ファイルはあったけどどこかでエラーが起きた時を省く
        try:
            if player_data_dict["today"] == "Error":
                main_content = {
                    "content": f"{now.month}月{now.day}日 : データなし"
                }
                requests.post(break_webhook_url, main_content) #結果をウェブフックに投稿
                player_data_dict.clear() #翌日のためにplayer_data_dictを空にする
                with open("player_data.json", mode="w") as f:
                    player_data_json = json.dumps(player_data_dict, indent=4)
                    f.write(player_data_json)
                return

        #KeyErrorが出れば正常。処理を行う
        except KeyError:
            daily_player_data_dict = {}
            url = f"https://w4.minecraftserver.jp/api/ranking?type=break&offset=0&lim=20&duration=daily" #今度は全MCIDではなく上位20人のみ
            try:
                res = requests.get(url) #さっきと同じ
                res.raise_for_status()
                player_data_every_day_dict = ast.literal_eval(str(bs4.BeautifulSoup(res.text, "html.parser")))
                player_data_every_day_list = player_data_every_day_dict["ranks"]
                for player_data_every_day in player_data_every_day_list:
                    mcid = player_data_every_day["player"]["name"]
                    try:
                        player_data = player_data_dict[mcid]
                    except KeyError:
                        player_data = {}
                    hour_break = {}
                    for hour in range(24): #0～23の24回ループ
                        try:
                            break_amount = player_data[f"{hour}"] #時間毎の整地量を取得
                        except KeyError:
                            break_amount = 0 #整地されていなければ0を入力
                        hour_break[f"{hour}"] = break_amount
                    raw_data = player_data_every_day["data"]["raw_data"] #23:58分の整地量を取得
                    hour_break["23_58"] = int(raw_data) #整地量を記録
                    daily_player_data_dict[mcid] = hour_break #MCIDごとに時間毎の整地量を記録
                today = datetime.date.today().strftime(r"%Y%m%d") #str型の今日の日付
                with open(f"{today}.json", mode="w") as f:
                    daily_player_data_json = json.dumps(daily_player_data_dict, indent=4)
                    f.write(daily_player_data_json) #記録したデータをJSON形式で書き込み
                
                player_data_dict.clear() #翌日のためにplayer_data_dictを空にする
                with open("player_data.json", mode="w") as f:
                    player_data_json = json.dumps(player_data_dict, indent=4)
                    f.write(player_data_json)
                pyxl() #データをエクセルに入力する関数
                draw_graph() #グラフを作成する関数
                push_files() #ファイルをウェブフックに投稿する関数

            #ウェブページに不調があったときの処理
            except requests.exceptions.HTTPError:
                main_content = {
                    "content": f"{now.month}月{now.day}日 : データなし"
                }
                requests.post(break_webhook_url, main_content) #結果をウェブフックに投稿
                player_data_dict.clear() #翌日のためにplayer_data_dictを空にする
                with open("player_data.json", mode="w") as f:
                    player_data_json = json.dumps(player_data_dict, indent=4)
                return
    except:
        unexpected_error() #こちらでは考えつかないエラーが起きたときunexpected_error()に回す。


login() #login()を実行。プログラム起動時に実行される。
schedule.every().hour.at(":01").do(job_every_hour) #job_every_hour()を毎時1分に実行
schedule.every().day.at("23:58").do(job_every_day) #job_every_every()を毎日23:58に実行

#ちょっとは動作軽くしたいじゃん？一秒間に何回も確認してたら重くなるやん？ってことで1秒休み
while True:
    schedule.run_pending()
    time.sleep(1)